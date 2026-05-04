"""
SHEtoken Pipeline — Excel Backup Writer
=========================================
Generates a local Excel backup of all WEI data.
Same 5 tabs as the Google Sheet — use as offline backup
or to share directly with partners without a Google account.

Usage:
    python data/write_to_excel.py

    # Or as part of full pipeline:
    python run_pipeline.py --excel

Output:
    data/output/SHEtoken_WEI_2025.xlsx

© 2026 SHE Foundation. Licensed under MIT.
"""

import os
import sys
import pandas as pd
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
from config import OUTPUT_DIR, BASELINE_YEAR


# ── BRAND COLOURS (openpyxl uses ARGB hex, no #) ─────────────────────────────

BERRY       = "FF6D2E46"
GOLD        = "FFC9A84C"
GOLD_LIGHT  = "FFF9EEF2"
CREAM       = "FFECE2D0"
DARK        = "FF1A0A12"
WHITE       = "FFFFFFFF"
GREEN_LIGHT = "FFD4EDDA"
RED_LIGHT   = "FFFFD7D7"
AMBER_LIGHT = "FFFFF3CD"
GREY_LIGHT  = "FFF5F5F5"
GREY_MID    = "FFE0E0E0"


# ── STYLE HELPERS ─────────────────────────────────────────────────────────────

def berry_fill():
    return PatternFill("solid", fgColor=BERRY)

def gold_fill():
    return PatternFill("solid", fgColor=GOLD)

def gold_light_fill():
    return PatternFill("solid", fgColor=GOLD_LIGHT)

def grey_fill():
    return PatternFill("solid", fgColor=GREY_LIGHT)

def dark_fill():
    return PatternFill("solid", fgColor=DARK)

def score_fill(score):
    """Return fill colour based on WEI score."""
    try:
        s = float(score)
        if s >= 70:   return PatternFill("solid", fgColor=GREEN_LIGHT)
        if s >= 45:   return PatternFill("solid", fgColor=GOLD_LIGHT)
        if s >= 20:   return PatternFill("solid", fgColor=AMBER_LIGHT)
        return             PatternFill("solid", fgColor=RED_LIGHT)
    except (TypeError, ValueError):
        return None

def title_font(size=12):
    return Font(name="Arial", bold=True, color=GOLD, size=size)

def header_font():
    return Font(name="Arial", bold=True, color=WHITE, size=10)

def body_font(bold=False, size=10):
    return Font(name="Arial", bold=bold, size=size)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=False)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=False)

def thin_border():
    s = Side(style="thin", color="FFD0D0D0")
    return Border(bottom=s)

def set_col_widths(ws, widths):
    """widths: list of (col_letter_or_num, width)"""
    for col, w in widths:
        if isinstance(col, int):
            col = get_column_letter(col)
        ws.column_dimensions[col].width = w


def write_banner(ws, row, text, merge_to_col, size=12):
    """Write a full-width berry banner row."""
    ws.row_dimensions[row].height = 22
    cell = ws.cell(row=row, column=1, value=text)
    cell.font      = title_font(size)
    cell.fill      = berry_fill()
    cell.alignment = center()
    ws.merge_cells(
        start_row=row, start_column=1,
        end_row=row,   end_column=merge_to_col
    )


def write_headers(ws, row, headers):
    """Write a header row with berry background."""
    ws.row_dimensions[row].height = 18
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font      = header_font()
        cell.fill      = berry_fill()
        cell.alignment = center()


def write_data_row(ws, row, values, alternate=False):
    """Write a data row with optional alternating background."""
    fill = grey_fill() if alternate else None
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font      = body_font()
        cell.alignment = left() if col_idx <= 2 else center()
        if fill:
            cell.fill = fill


# ── TAB 1: WEI SCORES (GLOBAL) ────────────────────────────────────────────────

def write_global_scores(wb, df, year):
    ws = wb.active
    ws.title = "WEI Scores"
    ws.freeze_panes = "A3"

    write_banner(ws, 1,
        f"SHEtoken — Women's Empowerment Index {year}  |  Global Country Scores  |  shetoken.org",
        17, size=11
    )

    headers = [
        "Rank", "Country", "ISO", "Region", "Tier",
        "Population (M)", "WEI Score", "Empowerment",
        "Education", "Economic", "Health", "Crime Penalty",
        "Ticker", "Data Source", "Verified", "Year", "Notes"
    ]
    write_headers(ws, 2, headers)

    for i, (_, r) in enumerate(df.iterrows()):
        row = i + 3
        values = [
            r.get("rank", ""),
            r.get("country", ""),
            r.get("iso_code", ""),
            r.get("region", ""),
            r.get("tier", ""),
            r.get("population_millions", ""),
            r.get("wei_score", ""),
            r.get("empowerment_score", ""),
            r.get("education_score", ""),
            r.get("economic_score", ""),
            r.get("health_score", ""),
            r.get("crime_penalty_score", ""),
            r.get("ticker", ""),
            r.get("data_source", ""),
            r.get("verified", ""),
            r.get("year", year),
            r.get("notes", ""),
        ]
        write_data_row(ws, row, values, alternate=(i % 2 == 0))

        # Colour the WEI score cell
        fill = score_fill(r.get("wei_score"))
        if fill:
            ws.cell(row=row, column=7).fill = fill
            ws.cell(row=row, column=7).font = body_font(bold=True)

    set_col_widths(ws, [
        ("A", 6), ("B", 22), ("C", 7), ("D", 14), ("E", 6),
        ("F", 13), ("G", 10), ("H", 13), ("I", 11), ("J", 11),
        ("K", 9), ("L", 14), ("M", 12), ("N", 14), ("O", 10),
        ("P", 7), ("Q", 20),
    ])


# ── TAB 2: INDIA STATES ───────────────────────────────────────────────────────

def write_india_states(wb, df, year):
    ws = wb.create_sheet("India States")
    ws.freeze_panes = "A3"

    write_banner(ws, 1,
        f"SHEtoken — India State WEI Scores {year}  |  shetoken.org",
        17, size=11
    )

    headers = [
        "Rank", "State", "Code", "Ticker", "Region",
        "Population (M)", "WEI Score", "vs Last Year", "HOT",
        "Empowerment", "Education", "Economic", "Health", "Crime Penalty",
        "Verified", "Year", "Key Programs"
    ]
    write_headers(ws, 2, headers)

    for i, (_, r) in enumerate(df.iterrows()):
        row = i + 3
        change = r.get("change", "")
        change_str = f"+{change}" if isinstance(change, (int, float)) and change > 0 else str(change)
        hot = "HOT" if str(r.get("hot", "")).lower() == "true" else ""

        values = [
            r.get("rank", ""),
            r.get("state", ""),
            r.get("state_code", ""),
            r.get("ticker", ""),
            r.get("region", ""),
            r.get("population_millions", ""),
            r.get("wei_score", ""),
            change_str,
            hot,
            r.get("empowerment_score", ""),
            r.get("education_score", ""),
            r.get("economic_score", ""),
            r.get("health_score", ""),
            r.get("crime_penalty_score", ""),
            r.get("verified", ""),
            r.get("year", year),
            str(r.get("key_programs", ""))[:120],
        ]
        write_data_row(ws, row, values, alternate=(i % 2 == 0))

        # Highlight HOT states in gold
        if hot == "HOT":
            for col in range(1, 10):
                ws.cell(row=row, column=col).fill = gold_light_fill()
                ws.cell(row=row, column=col).font = body_font(bold=True)

        # Colour change cell
        try:
            chg = float(r.get("change", 0) or 0)
            change_cell = ws.cell(row=row, column=8)
            if chg > 0:
                change_cell.fill = PatternFill("solid", fgColor=GREEN_LIGHT)
                change_cell.font = Font(name="Arial", bold=True, color="FF1A6B34", size=10)
            elif chg < 0:
                change_cell.fill = PatternFill("solid", fgColor=RED_LIGHT)
                change_cell.font = Font(name="Arial", bold=True, color="FF8B0000", size=10)
        except (TypeError, ValueError):
            pass

        # WEI score colour
        fill = score_fill(r.get("wei_score"))
        if fill:
            ws.cell(row=row, column=7).fill = fill
            ws.cell(row=row, column=7).font = body_font(bold=True)

    set_col_widths(ws, [
        ("A", 6), ("B", 20), ("C", 7), ("D", 12), ("E", 10),
        ("F", 13), ("G", 10), ("H", 12), ("I", 8),
        ("J", 13), ("K", 11), ("L", 11), ("M", 9), ("N", 14),
        ("O", 10), ("P", 7), ("Q", 50),
    ])


# ── TAB 3: LEADERBOARD ────────────────────────────────────────────────────────

def write_leaderboard(wb, global_df, india_df, year):
    ws = wb.create_sheet("Leaderboard")
    row = 1

    write_banner(ws, row, f"SHEtoken — WEI Leaderboard {year}  |  shetoken.org", 7)
    row += 2

    # Top 10 countries
    sub_banner = ws.cell(row=row, column=1,
                          value="TOP 10 COUNTRIES — Highest WEI Score")
    sub_banner.font = Font(name="Arial", bold=True, color=GOLD, size=11)
    sub_banner.fill = berry_fill()
    sub_banner.alignment = center()
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    row += 1

    top10_headers = ["Rank", "Country", "Ticker", "WEI Score",
                     "Empowerment", "Education", "Crime Penalty"]
    write_headers(ws, row, top10_headers)
    row += 1

    top10 = global_df.nlargest(10, "wei_score")
    for i, (_, r) in enumerate(top10.iterrows()):
        vals = [
            r.get("rank"), r.get("country"), r.get("ticker"),
            r.get("wei_score"), r.get("empowerment_score"),
            r.get("education_score"), r.get("crime_penalty_score"),
        ]
        write_data_row(ws, row, vals, alternate=(i % 2 == 0))
        ws.cell(row=row, column=4).fill = PatternFill("solid", fgColor=GREEN_LIGHT)
        ws.cell(row=row, column=4).font = body_font(bold=True)
        row += 1

    row += 1

    # India State leaderboard
    sub2 = ws.cell(row=row, column=1, value="INDIA STATE LEADERBOARD")
    sub2.font = Font(name="Arial", bold=True, color=GOLD, size=11)
    sub2.fill = berry_fill()
    sub2.alignment = center()
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    row += 1

    india_headers = ["Rank", "State", "Ticker", "WEI Score",
                     "vs Last Year", "HOT", "Key Program Driver"]
    write_headers(ws, row, india_headers)
    row += 1

    india_sorted = india_df.sort_values("wei_score", ascending=False)
    for i, (_, r) in enumerate(india_sorted.iterrows()):
        change = r.get("change", 0) or 0
        try:
            change_str = f"+{float(change)}" if float(change) > 0 else str(change)
        except (TypeError, ValueError):
            change_str = str(change)
        hot = "HOT" if str(r.get("hot", "")).lower() == "true" else ""
        programs = str(r.get("key_programs", ""))
        driver = programs.split("—")[0].strip()[:45] if "—" in programs else programs[:45]

        vals = [r.get("rank"), r.get("state"), r.get("ticker"),
                r.get("wei_score"), change_str, hot, driver]
        write_data_row(ws, row, vals, alternate=(i % 2 == 0))

        if hot == "HOT":
            for col in range(1, 8):
                ws.cell(row=row, column=col).fill = gold_light_fill()
                ws.cell(row=row, column=col).font = body_font(bold=True)
        row += 1

    set_col_widths(ws, [
        ("A", 7), ("B", 22), ("C", 12), ("D", 11),
        ("E", 13), ("F", 8), ("G", 48),
    ])


# ── TAB 4: SUMMARY ────────────────────────────────────────────────────────────

def write_summary(wb, global_df, india_df, year):
    ws = wb.create_sheet("Summary Dashboard")

    global_wei = round(
        (global_df["wei_score"] * global_df["population_millions"]).sum()
        / global_df["population_millions"].sum(), 1
    )

    tier_counts = global_df["tier"].value_counts().to_dict()
    highest     = global_df.loc[global_df["wei_score"].idxmax()]
    lowest      = global_df.loc[global_df["wei_score"].idxmin()]
    india_row   = global_df[global_df["iso_code"] == "IND"]
    india_wei   = india_row["wei_score"].values[0] if len(india_row) else "N/A"

    top_state  = india_df.loc[india_df["wei_score"].idxmax()] if len(india_df) else None
    fast_state = india_df.loc[india_df["change"].idxmax()] if (
        len(india_df) and "change" in india_df.columns) else None

    now = datetime.now().strftime("%B %d, %Y  %H:%M UTC")

    sections = [
        ("GLOBAL WEI INDEX", [
            ("Global WEI Score (population-weighted)",
             global_wei,
             "=AVERAGE(\'WEI Scores\'!G3:G85)"),
            ("Countries Scored",      len(global_df), ""),
            ("India WEI Score ($SHE-IND)", india_wei, ""),
            ("Last Updated",          now,            ""),
        ]),
        ("TIER BREAKDOWN", [
            ("Tier 1 — WEI 70+ (high performing)",  tier_counts.get(1, 0), "Nordic, Anglosphere"),
            ("Tier 2 — WEI 45–69 (mid performing)", tier_counts.get(2, 0), "India, Brazil, Kenya"),
            ("Tier 3 — WEI 20–44 (developing)",     tier_counts.get(3, 0), "Pakistan, Nigeria"),
            ("Tier 4 — WEI <20 (crisis level)",     tier_counts.get(4, 0), "Yemen, Afghanistan"),
        ]),
        ("TOP & BOTTOM", [
            ("Highest WEI",
             f"{highest['country']} ({highest['wei_score']})", ""),
            ("Lowest WEI",
             f"{lowest['country']} ({lowest['wei_score']})", ""),
        ]),
        ("INDIA STATES", [
            ("States Scored",   len(india_df), ""),
            ("Highest WEI State",
             f"{top_state['state']} ({top_state['wei_score']})" if top_state is not None else "", ""),
            ("Fastest Improving",
             f"{fast_state['state']} (+{fast_state.get('change','')}pts)" if fast_state is not None else "", ""),
        ]),
        ("TOKEN MECHANICS", [
            ("+1 WEI point → 10M SHE minted", "to WEI Impact Fund", ""),
            ("−1 WEI point → 10M SHE burned", "permanently removed from supply", ""),
            ("Crime spike >15%",               "Crisis Trigger (DAO vote)", ""),
        ]),
        ("LINKS", [
            ("Website",    "shetoken.org",              ""),
            ("GitHub",     "github.com/shetoken",       ""),
            ("Whitepaper", "shetoken.org/whitepaper",   ""),
            ("Contact",    "contact@shetoken.org",      ""),
        ]),
    ]

    write_banner(ws, 1,
        f"SHEtoken — WEI Summary Dashboard {year}  |  shetoken.org",
        3, size=12
    )

    row = 3
    for section_title, rows_data in sections:
        # Section header
        cell = ws.cell(row=row, column=1, value=section_title)
        cell.font = Font(name="Arial", bold=True, color=GOLD, size=10)
        cell.fill = berry_fill()
        cell.alignment = left()
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=3)
        row += 1

        for label, value, note in rows_data:
            ws.cell(row=row, column=1, value=label).font  = body_font()
            val_cell = ws.cell(row=row, column=2, value=value)
            val_cell.font = body_font(bold=True)
            val_cell.alignment = center()

            # Highlight global WEI score
            if label.startswith("Global WEI Score"):
                val_cell.fill = gold_light_fill()
                val_cell.font = Font(name="Arial", bold=True, size=14)

            if note:
                ws.cell(row=row, column=3, value=note).font = Font(
                    name="Arial", color="FF888888", size=9, italic=True)
            row += 1

        row += 1

    set_col_widths(ws, [("A", 42), ("B", 32), ("C", 38)])


# ── TAB 5: METHODOLOGY ───────────────────────────────────────────────────────

def write_methodology(wb):
    ws = wb.create_sheet("Methodology")

    write_banner(ws, 1,
        "SHEtoken — WEI Methodology Reference  |  shetoken.org",
        4, size=11
    )

    sections = [
        ("THE WEI FORMULA", [
            ["WEI = (Empowerment × 0.25) + (Education × 0.20) + (Economic × 0.20) + (Health × 0.15) − (Crime Penalty × 0.20)", "", "", ""],
            ["All pillar scores normalised 0–100. Crime Penalty subtracted — higher crime = lower WEI.", "", "", ""],
        ]),
        ("PILLAR WEIGHTS & SOURCES", [
            ["Pillar", "Weight", "Key Indicators", "Primary Source"],
            ["Empowerment",   "25%", "Women in parliament, ministerial roles, legal rights", "IPU Parline, UN Women"],
            ["Education",     "20%", "Female literacy rate, school enrollment",             "UNESCO UIS, World Bank"],
            ["Economic",      "20%", "Gender wage gap, female LFPR, bank accounts",          "ILO, World Bank"],
            ["Health",        "15%", "Maternal mortality, life expectancy",                  "WHO GHO, World Bank"],
            ["Crime Penalty", "20%", "Rape rate, domestic violence, femicide",               "UNODC, NCRB (India)"],
        ]),
        ("COUNTRY TIERS", [
            ["Tier", "WEI Range", "Examples", "Population Weight"],
            ["Tier 1 — High performing",  "70–100", "Iceland, Norway, NZ",     "1.0×"],
            ["Tier 2 — Mid performing",   "45–69",  "India, Brazil, Kenya",    "1.0×"],
            ["Tier 3 — Developing",       "20–44",  "Pakistan, Nigeria",       "0.8×"],
            ["Tier 4 — Crisis level",     "0–19",   "Yemen, Afghanistan",      "0.6×"],
        ]),
        ("TOKEN MECHANICS", [
            ["Global WEI +1 point", "→ 10,000,000 SHE tokens minted", "Added to WEI Impact Fund", ""],
            ["Global WEI −1 point", "→ 10,000,000 SHE tokens burned",  "Permanently removed",     ""],
            ["Crime spike >15%",    "→ Crisis Trigger",                  "DAO governance vote",    ""],
        ]),
        ("TRANSPARENCY", [
            ["Open-source methodology", "github.com/shetoken",                          "", ""],
            ["Challenge a score",       "GitHub Issue: label wei-challenge",             "", ""],
            ["Annual review window",    "30-day public review after each publication",  "", ""],
            ["Full methodology",        "wei-index/methodology.md",                     "", ""],
        ]),
    ]

    row = 3
    for section_title, rows_data in sections:
        cell = ws.cell(row=row, column=1, value=section_title)
        cell.font = Font(name="Arial", bold=True, color=GOLD, size=10)
        cell.fill = berry_fill()
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=4)
        row += 1

        for i, row_data in enumerate(rows_data):
            is_subheader = (section_title in (
                "PILLAR WEIGHTS & SOURCES", "COUNTRY TIERS", "TOKEN MECHANICS"
            ) and i == 0)

            for col_idx, val in enumerate(row_data[:4], 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                if is_subheader:
                    cell.font = header_font()
                    cell.fill = PatternFill("solid", fgColor="FF3D1829")
                else:
                    cell.font = body_font()
                    if i % 2 == 0:
                        cell.fill = grey_fill()
            row += 1

        row += 1

    set_col_widths(ws, [("A", 30), ("B", 36), ("C", 42), ("D", 22)])


# ── MAIN ──────────────────────────────────────────────────────────────────────

def write_excel(year=BASELINE_YEAR, output_dir=None):
    """
    Generate the Excel backup file.

    Args:
        year (int): Data year
        output_dir (Path): Output directory (defaults to data/output/)

    Returns:
        str: Path to generated Excel file
    """
    out_dir = output_dir or OUTPUT_DIR
    os.makedirs(out_dir, exist_ok=True)

    baseline_path = out_dir / f"baseline-{year}.csv"
    india_path    = out_dir / f"india-states-{year}.csv"

    if not baseline_path.exists():
        print(f"✗ {baseline_path} not found. Run generate_baseline.py first.")
        return None

    global_df = pd.read_csv(baseline_path, comment="#")
    india_df  = pd.read_csv(india_path, comment="#") if india_path.exists() else pd.DataFrame()

    print(f"  Loaded: {len(global_df)} countries, {len(india_df)} India states")

    wb = Workbook()

    print("  Writing tabs...")
    write_global_scores(wb, global_df, year)
    print("    ✓ WEI Scores")
    write_india_states(wb, india_df, year)
    print("    ✓ India States")
    write_leaderboard(wb, global_df, india_df, year)
    print("    ✓ Leaderboard")
    write_summary(wb, global_df, india_df, year)
    print("    ✓ Summary Dashboard")
    write_methodology(wb)
    print("    ✓ Methodology")

    out_path = out_dir / f"SHEtoken_WEI_{year}.xlsx"
    wb.save(str(out_path))

    print(f"\n✓ Excel backup saved: {out_path}")
    return str(out_path)


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(
        description="Generate Excel backup of WEI data"
    )
    parser.add_argument("--year", type=int, default=BASELINE_YEAR)
    args = parser.parse_args()

    print("SHEtoken — Excel Backup Generator")
    print("=" * 50)
    write_excel(year=args.year)
